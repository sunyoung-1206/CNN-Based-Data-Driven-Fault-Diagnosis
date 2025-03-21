# 기존 모델 로드 및 혼동 행렬 생성 코드

import tensorflow as tf
import numpy as np
import pandas as pd
import os
from PIL import Image
import openpyxl
import random
import matplotlib.pyplot as plt
import seaborn as sns
from sklearn.metrics import confusion_matrix

# 로깅 레벨 설정
os.environ['TF_CPP_MIN_LOG_LEVEL'] = '2'

# 경로 설정
IMAGESAVEURL_validation = "/home/potato/박선영/CNN-Based-Data-Driven-Fault-Diagnosis/generated_data(2000)/test_images"

# Excel 파일 로드
validation_wb = openpyxl.load_workbook('/home/potato/박선영/CNN-Based-Data-Driven-Fault-Diagnosis/generated_data(2000)/validationImageList.xlsx')
validation_sheet = validation_wb.active

# 테스트 데이터 로드 함수
def load_test_data(sheet, img_dir, num_samples):
    images = []
    labels = []
    skipped = 0
    processed = 0

    print(f"테스트 데이터 처리 시작...")

    # 테스트 데이터 수집
    data_entries = []
    for i in range(num_samples):
        try:
            imagename = sheet.cell(i+1, 1).value
            label_index = int(sheet.cell(i+1, 2).value)
            if imagename is None or label_index < 0 or label_index >= 10:
                skipped += 1
                continue

            image_path = os.path.join(img_dir, imagename)
            if not os.path.exists(image_path):
                skipped += 1
                continue

            data_entries.append((image_path, label_index))
            processed += 1

            if (i+1) % 1000 == 0:
                print(f"테스트 데이터 처리 중... {i+1}/{num_samples} ({(i+1)/num_samples*100:.1f}%)")
        except Exception as e:
            skipped += 1
            continue

    # 데이터 셔플
    random.shuffle(data_entries)

    # 셔플된 데이터 로드
    for img_path, label in data_entries:
        try:
            image = Image.open(img_path)
            # 이미지 크기 확인 및 조정
            width, height = image.size
            if width != 64 or height != 64:
                image = image.resize((64, 64))

            # 이미지 전처리
            imagedata = np.array(image.convert('L'), dtype='float32') / 255.0

            # 결과 저장
            images.append(imagedata)

            # 원-핫 인코딩
            one_hot = np.zeros(10, dtype=np.float32)
            one_hot[label] = 1.0
            labels.append(one_hot)
        except Exception as e:
            print(f"이미지 로드 중 오류 발생: {img_path}, {str(e)}")

    # 배열로 변환
    images = np.array(images)
    labels = np.array(labels)

    print(f"테스트 데이터 처리 완료")
    print(f"처리된 이미지: {processed}/{num_samples}, 스킵된 이미지: {skipped}/{num_samples}")

    return images, labels

# 테스트 데이터 로드 (16000개 샘플)
print("테스트 데이터 로드 중...")
x_test, y_test = load_test_data(validation_sheet, IMAGESAVEURL_validation, 16000)

# 데이터셋 생성
batch_size = 32
test_dataset = tf.data.Dataset.from_tensor_slices((
    np.expand_dims(x_test, -1),  # 채널 차원 추가
    y_test
)).batch(batch_size)

try:
    # 저장된 모델 로드
    model_path = '최고의_cnn_model.h5'  # 모델 파일 경로
    print(f"\n'{model_path}' 모델 로드 중...")
    model = tf.keras.models.load_model(model_path)

    # 모델 요약 정보 출력
    model.summary()

    # 테스트 데이터로 모델 평가
    print("\n모델 평가 중...")
    test_loss, test_accuracy = model.evaluate(test_dataset)
    print(f"테스트 정확도: {test_accuracy:.4f}")

    # 전체 테스트 데이터에 대한 예측
    print("\n예측 수행 중...")
    predictions = model.predict(test_dataset)
    y_pred = np.argmax(predictions, axis=1)

    # 실제 라벨
    all_y_test = []
    for _, y_batch in test_dataset:
        all_y_test.extend(np.argmax(y_batch.numpy(), axis=1))
    y_true = np.array(all_y_test)

    # 클래스별 정확도 계산
    print("\n클래스별 정확도:")
    for c in range(10):
        mask = y_true == c
        if np.sum(mask) > 0:
            class_acc = np.mean(y_pred[mask] == c)
            print(f"  클래스 {c}: {class_acc:.4f} ({np.sum(mask)} 샘플)")

    # 혼동 행렬 생성 및 시각화
    print("\n혼동 행렬 생성 중...")
    cm = confusion_matrix(y_true, y_pred)

    # 퍼센트 값으로 변환 (0으로 나누는 문제 방지)
    cm_sum = cm.sum(axis=1, keepdims=True)
    cm_percentage = np.divide(cm.astype(float), cm_sum, where=(cm_sum != 0)) * 100
    cm_percentage = np.nan_to_num(cm_percentage)  # NaN을 0으로 변환

    # 클래스 라벨 정의
    labels = ["BF0.18", "BF0.36", "BF0.54", "OF0.18", "OF0.36", "OF0.54", "IF0.18", "IF0.36", "IF0.54", "NO"]

    plt.figure(figsize=(10, 8))

    # 히트맵 그리기
    sns.heatmap(cm_percentage, annot=True, fmt=".1f", cmap="gray_r",
                xticklabels=labels, yticklabels=labels, cbar=True)

    # 축 레이블 추가
    plt.xlabel("Predict label", fontsize=12)
    plt.ylabel("Actual label", fontsize=12)

    # 회전된 라벨 스타일 적용
    plt.xticks(rotation=45)
    plt.yticks(rotation=45)

    # 그래프 제목 설정
    plt.title("Confusion Matrix (%)", fontsize=14)

    # 그래프 저장
    plt.tight_layout()
    plt.savefig('confusion_matrix.png')
    print("혼동 행렬이 'confusion_matrix.png'로 저장되었습니다.")

    # 그래프 표시
    plt.show()

except Exception as e:
    print(f"오류 발생: {str(e)}")
