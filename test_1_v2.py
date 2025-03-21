# 0. 라이브러리 설치
# pip install tensorflow
# pip install numpy
# pip install pandas
# pip install openpyxl
# pip install matplotlib scikit-learn seaborn pillow

# 1. 라이브러리 임포트
# -*- coding: utf-8 -*-
"""
Original created on Tue May 1 15:31:04 2018
@author: zhaoyuzhi
Updated to modern TensorFlow API in 2025
"""

import tensorflow as tf
import numpy as np
import pandas as pd
import os
from PIL import Image
import openpyxl
import random
import time

# 로깅 레벨 설정
os.environ['TF_CPP_MIN_LOG_LEVEL'] = '2'

# 2. 경로 설정
# 경로 설정 (필요에 따라 수정)
IMAGESAVEURL_training = "/home/potato/박선영/CNN-Based-Data-Driven-Fault-Diagnosis/generated_data(2000)/train_images"
IMAGESAVEURL_validation = "/home/potato/박선영/CNN-Based-Data-Driven-Fault-Diagnosis/generated_data(2000)/test_images"

# 3. Excel 파일 로드
# Excel 파일 로드 (xlrd 대신 openpyxl 사용)
training_wb = openpyxl.load_workbook('/home/potato/박선영/CNN-Based-Data-Driven-Fault-Diagnosis/generated_data(2000)/trainingImageList.xlsx')
training_sheet = training_wb.active
validation_wb = openpyxl.load_workbook('/home/potato/박선영/CNN-Based-Data-Driven-Fault-Diagnosis/generated_data(2000)/validationImageList.xlsx')
validation_sheet = validation_wb.active

# 4. 데이터 배열 초기화
# 데이터 저장용 배열 초기화
train_num = 80000
test_num = 16000
x_data = np.zeros([train_num, 64, 64], dtype=np.float32)
y_data = np.zeros([train_num, 10], dtype=np.float32)
x_test = np.zeros([test_num, 64, 64], dtype=np.float32)
y_test = np.zeros([test_num, 10], dtype=np.float32)

# 디버깅 함수 추가 - 데이터 검증
def check_data_distribution(y_data):
    """데이터셋의 클래스 분포를 확인합니다."""
    label_counts = np.sum(y_data, axis=0)
    total_samples = np.sum(label_counts)

    print("\n데이터 분포 확인:")
    print("=" * 50)
    for i, count in enumerate(label_counts):
        percentage = (count / total_samples) * 100
        print(f"클래스 {i}: {int(count)} 샘플 ({percentage:.2f}%)")
    print("=" * 50)

    # 클래스 불균형 확인
    if max(label_counts) / (min(label_counts) + 1e-10) > 5:
        print("⚠ 경고: 심각한 클래스 불균형이 감지되었습니다.")
    return label_counts

# 디버깅 함수 추가 - 이미지 검증
def validate_images(x_data, sample_size=5):
    """이미지 데이터의 기본 통계를 확인합니다."""
    # 비어있는 이미지 확인
    empty_images = np.sum(np.all(x_data == 0, axis=(1, 2)))
    if empty_images > 0:
        print(f"⚠ 경고: {empty_images}개의 빈 이미지가 발견되었습니다.")

    # 픽셀 값 범위 확인
    min_val = np.min(x_data)
    max_val = np.max(x_data)
    mean_val = np.mean(x_data)
    std_val = np.std(x_data)

    print("\n이미지 통계:")
    print("=" * 50)
    print(f"최소값: {min_val}, 최대값: {max_val}")
    print(f"평균값: {mean_val:.4f}, 표준편차: {std_val:.4f}")

    if min_val < 0 or max_val > 1:
        print("⚠ 경고: 이미지 픽셀 값이 [0, 1] 범위를 벗어났습니다.")

    # 샘플 이미지의 평균과 표준편차
    indices = random.sample(range(len(x_data)), min(sample_size, len(x_data)))
    for i, idx in enumerate(indices):
        img = x_data[idx]
        print(f"샘플 {i+1}: 평균={np.mean(img):.4f}, 표준편차={np.std(img):.4f}")
    print("=" * 50)

# 5. 학습 데이터 처리
print("학습 데이터 처리 시작...")
start_time = time.time()
# 학습 데이터 처리
skipped_train = 0
processed_train = 0

for i in range(train_num):
    try:
        imagename = training_sheet.cell(i+1, 1).value  # 이미지 파일명
        # 파일명 검증 (None이면 스킵)
        if imagename is None:
            if i < 10:  # 처음 몇 개만 출력
                print(f"⚠ 스킵된 행: {i+1}, 파일명: {imagename}")
            skipped_train += 1
            continue

        image_path = os.path.join(IMAGESAVEURL_training, imagename)
        # 파일 존재 확인
        if not os.path.exists(image_path):
            if i < 10:  # 처음 몇 개만 출력
                print(f"⚠ 파일이 존재하지 않음: {image_path}")
            skipped_train += 1
            continue

        image = Image.open(image_path)

        # 이미지 사이즈 확인
        width, height = image.size
        if width != 64 or height != 64:
            print(f"⚠ 이미지 크기가 64x64가 아님: {width}x{height}")
            # 리사이징 적용
            image = image.resize((64, 64))

        # 이미지 전처리 (그레이스케일 변환 및 정규화)
        imagedata = np.array(image.convert('L'), dtype='float32') / 255.0
        x_data[i, :, :] = imagedata

        # 10개 클래스에 맞춰 원-핫 벡터 생성
        label_index = int(training_sheet.cell(i+1, 2).value)  # 예: 0~9 사이의 값
        if label_index < 0 or label_index >= 10:
            print(f"⚠ 유효하지 않은 레이블: {label_index}, 행: {i+1}")
            skipped_train += 1
            continue

        y_data[i, label_index] = 1  # 해당 위치만 1로 설정
        processed_train += 1

        # 진행상황 표시
        if (i+1) % 5000 == 0:
            print(f"학습 데이터 처리 중... {i+1}/{train_num} ({(i+1)/train_num*100:.1f}%)")

    except Exception as e:
        if i < 10:  # 처음 몇 개만 출력
            print(f"처리 중 오류 발생 (학습 이미지 {i+1}): {str(e)}")
        skipped_train += 1
        continue

print(f"학습 데이터 처리 완료 ({time.time() - start_time:.2f}초)")
print(f"처리된 이미지: {processed_train}/{train_num}, 스킵된 이미지: {skipped_train}/{train_num}")

# 데이터 확인
if processed_train > 0:
    # 실제로 사용할 데이터만 선택
    valid_indices = np.where(np.sum(y_data, axis=1) > 0)[0]
    if len(valid_indices) < train_num:
        print(f"유효한 학습 데이터로 배열 크기 조정: {len(valid_indices)}/{train_num}")
        x_data = x_data[valid_indices]
        y_data = y_data[valid_indices]

# 데이터 분포와 이미지 품질 확인
train_label_counts = check_data_distribution(y_data)
validate_images(x_data)

# 6. 테스트 데이터 처리
print("\n테스트 데이터 처리 시작...")
start_time = time.time()
skipped_test = 0
processed_test = 0

for i in range(test_num):
    try:
        # 이미지 불러오기
        imagename = validation_sheet.cell(i+1, 1).value
        if imagename is None:
            if i < 10:  # 처음 몇 개만 출력
                print(f"⚠ 스킵된 행: {i+1}, 파일명: {imagename}")
            skipped_test += 1
            continue

        image_path = os.path.join(IMAGESAVEURL_validation, imagename)
        # 파일 존재 확인
        if not os.path.exists(image_path):
            if i < 10:  # 처음 몇 개만 출력
                print(f"⚠ 파일이 존재하지 않음: {image_path}")
            skipped_test += 1
            continue

        image = Image.open(image_path)

        # 이미지 사이즈 확인
        width, height = image.size
        if width != 64 or height != 64:
            print(f"⚠ 이미지 크기가 64x64가 아님: {width}x{height}")
            # 리사이징 적용
            image = image.resize((64, 64))

        # 이미지를 그레이스케일로 변환하고 정규화
        imagedata = np.array(image.convert('L'), dtype='float32') / 255.0

        # 데이터 저장
        x_test[i, :, :] = imagedata

        # 10개 클래스에 맞춰 원-핫 벡터 생성
        label_index = int(validation_sheet.cell(i+1, 2).value)  # 예: 0~9 사이의 값
        if label_index < 0 or label_index >= 10:
            print(f"⚠ 유효하지 않은 레이블: {label_index}, 행: {i+1}")
            skipped_test += 1
            continue

        y_test[i, label_index] = 1  # 해당 위치만 1로 설정
        processed_test += 1

        # 진행상황 표시
        if (i+1) % 1000 == 0:
            print(f"테스트 데이터 처리 중... {i+1}/{test_num} ({(i+1)/test_num*100:.1f}%)")

    except Exception as e:
        if i < 10:  # 처음 몇 개만 출력
            print(f"처리 중 오류 발생 (테스트 이미지 {i+1}): {str(e)}")
        skipped_test += 1
        continue

print(f"테스트 데이터 처리 완료 ({time.time() - start_time:.2f}초)")
print(f"처리된 이미지: {processed_test}/{test_num}, 스킵된 이미지: {skipped_test}/{test_num}")

# 데이터 확인
if processed_test > 0:
    # 실제로 사용할 데이터만 선택
    valid_indices = np.where(np.sum(y_test, axis=1) > 0)[0]
    if len(valid_indices) < test_num:
        print(f"유효한 테스트 데이터로 배열 크기 조정: {len(valid_indices)}/{test_num}")
        x_test = x_test[valid_indices]
        y_test = y_test[valid_indices]

# 데이터 분포와 이미지 품질 확인
test_label_counts = check_data_distribution(y_test)
validate_images(x_test)

# 7. CNN 모델 정의 함수
# 모델 정의 (하이퍼파라미터 유지)
def create_cnn_model():
    # 함수형 API 사용
    inputs = tf.keras.Input(shape=(64, 64, 1))

    # 첫 번째 합성곱 레이어 및 풀링
    x = tf.keras.layers.Conv2D(32, (5, 5), padding='same', activation='relu')(inputs)
    x = tf.keras.layers.MaxPooling2D((2, 2))(x)

    # 두 번째 합성곱 레이어 및 풀링
    x = tf.keras.layers.Conv2D(64, (3, 3), padding='same', activation='relu')(x)
    x = tf.keras.layers.MaxPooling2D((2, 2))(x)

    # 세 번째 합성곱 레이어 및 풀링
    x = tf.keras.layers.Conv2D(128, (3, 3), padding='same', activation='relu')(x)
    x = tf.keras.layers.MaxPooling2D((2, 2))(x)

    # 네 번째 합성곱 레이어 및 풀링
    x = tf.keras.layers.Conv2D(256, (3, 3), padding='same', activation='relu')(x)
    x = tf.keras.layers.MaxPooling2D((2, 2))(x)

    # 플래튼 레이어
    x = tf.keras.layers.Flatten()(x)

    # 완전 연결 레이어 1
    x = tf.keras.layers.Dense(2560, activation='relu')(x)
    x = tf.keras.layers.Dropout(0.6)(x)  # keep_prob 0.4 -> dropout 0.6

    # 완전 연결 레이어 2 (출력 레이어)
    outputs = tf.keras.layers.Dense(10, activation='softmax')(x)

    # 모델 생성
    model = tf.keras.Model(inputs=inputs, outputs=outputs)

    return model

# 8. 모델 생성 및 컴파일 (논문 하이퍼파라미터 유지)
# 모델 생성 및 컴파일
model = create_cnn_model()
model.compile(
    optimizer=tf.keras.optimizers.Adam(learning_rate=1e-6),  # 논문 그대로 유지
    loss='categorical_crossentropy',
    metrics=['accuracy']
)
model.summary()

# 9. 데이터셋 생성
# TensorFlow 데이터셋 API 사용 (배치 사이즈 유지)
batch_size = 16  # 논문 그대로 유지
train_dataset = tf.data.Dataset.from_tensor_slices((
    np.expand_dims(x_data, -1),  # 채널 차원 추가
    y_data
)).batch(batch_size)

# 테스트용 데이터셋 생성
test_dataset = tf.data.Dataset.from_tensor_slices((
    np.expand_dims(x_test, -1),  # 채널 차원 추가
    y_test
)).batch(batch_size)

# 학습 성능 모니터링을 위한 콜백 추가
class TrainingMonitor(tf.keras.callbacks.Callback):
    def __init__(self, test_data, test_interval=1):
        super(TrainingMonitor, self).__init__()
        self.test_data = test_data
        self.test_interval = test_interval
        self.history = {'val_accuracy': [], 'val_loss': []}

    def on_epoch_end(self, epoch, logs=None):
        if (epoch + 1) % self.test_interval == 0:
            val_loss, val_acc = self.model.evaluate(self.test_data, verbose=0)
            self.history['val_accuracy'].append(val_acc)
            self.history['val_loss'].append(val_loss)
            print(f"\n에폭 {epoch+1} 검증 성능 - 정확도: {val_acc:.4f}, 손실: {val_loss:.4f}")

# 모니터링 콜백 생성
monitor = TrainingMonitor(test_dataset, test_interval=1)

# 10. 모델 학습 (에폭 수 유지)
print("\n모델 학습 시작...")
epochs = 12  # 논문 그대로 유지
history = model.fit(
    train_dataset,
    epochs=epochs,
    callbacks=[monitor],
    verbose=1
)

# 11. 모델 평가
# 테스트 데이터로 모델 평가
print("\n최종 모델 평가...")
test_loss, test_accuracy = model.evaluate(test_dataset)
print(f"테스트 정확도: {test_accuracy:.4f}")

# 12. 예측 결과 저장
# 예측 결과 저장
predictions = model.predict(test_dataset)
print(f"예측 형태: {predictions.shape}")

# 13. 모델 저장
# 모델 저장
model.save('cnn_model.h5')
print("모델이 cnn_model.h5로 저장되었습니다.")
print("\n실행이 모두 완료되었습니다.")
