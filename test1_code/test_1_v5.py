#!/usr/bin/env python3
"""
test1_v5_no_chunking.py - 메모리 제한 및 청킹 제거, 미니배치 16으로 변경
"""

import tensorflow as tf
import numpy as np
import pandas as pd
import os
from PIL import Image
import openpyxl
import random
import time
import matplotlib.pyplot as plt
import seaborn as sns
from sklearn.metrics import confusion_matrix
import gc

# 로깅 레벨 설정
os.environ['TF_CPP_MIN_LOG_LEVEL'] = '2'

# 경로 설정
IMAGESAVEURL_training = "/home/potato/박선영/CNN-Based-Data-Driven-Fault-Diagnosis/test1_generated_data/train_images"
IMAGESAVEURL_validation = "/home/potato/박선영/CNN-Based-Data-Driven-Fault-Diagnosis/test1_generated_data/test_images"

# Excel 파일 로드
training_wb = openpyxl.load_workbook('/home/potato/박선영/CNN-Based-Data-Driven-Fault-Diagnosis/test1_generated_data/trainingImageList.xlsx')
training_sheet = training_wb.active
validation_wb = openpyxl.load_workbook('/home/potato/박선영/CNN-Based-Data-Driven-Fault-Diagnosis/test1_generated_data/validationImageList.xlsx')
validation_sheet = validation_wb.active

# 데이터 배열 초기화 - 상단에 설정하여 쉽게 변경 가능
CONFIG = {
    'train_num': 80000,     # 원래 값 유지
    'test_num': 16000,      # 원래 값 유지
    'num_classes': 10,      # 분류할 클래스 수
    'model_prefix': 'test1_v5',  # 저장 파일 접두사
    'class_labels': ["BF0.18", "BF0.36", "BF0.54", "OF0.18", "OF0.36", "OF0.54", "IF0.18", "IF0.36", "IF0.54", "NO"],  # 혼동 행렬에 표시할 클래스 라벨
    'batch_size': 16,       # 미니배치 크기를 16으로 변경
}

# 전체 훈련 데이터 로드
def load_training_data():
    print("훈련 데이터 로드 중...")
    images = []
    labels = []
    skipped = 0

    for i in range(CONFIG['train_num']):
        try:
            imagename = training_sheet.cell(i+1, 1).value
            label_index = int(training_sheet.cell(i+1, 2).value)

            if imagename is None or label_index < 0 or label_index >= CONFIG['num_classes']:
                skipped += 1
                continue

            image_path = os.path.join(IMAGESAVEURL_training, imagename)
            if not os.path.exists(image_path):
                skipped += 1
                continue

            # 이미지 로드
            image = Image.open(image_path)

            # 이미지 크기 확인 및 조정
            width, height = image.size
            if width != 64 or height != 64:
                image = image.resize((64, 64))

            # 이미지 전처리
            imagedata = np.array(image.convert('L'), dtype='float32') / 255.0
            images.append(imagedata)

            # 원-핫 인코딩
            one_hot = np.zeros(CONFIG['num_classes'], dtype=np.float32)
            one_hot[label_index] = 1.0
            labels.append(one_hot)

            if (i+1) % 10000 == 0:
                print(f"처리 중... {i+1}/{CONFIG['train_num']} ({(i+1)/CONFIG['train_num']*100:.1f}%)")

        except Exception as e:
            skipped += 1
            print(f"오류 발생: {e}")
            continue

    print(f"수집된 훈련 데이터: {len(images)}, 스킵됨: {skipped}")

    # 배열로 변환
    x_train = np.array(images)
    y_train = np.array(labels)

    # 채널 차원 추가
    x_train = np.expand_dims(x_train, -1)

    return x_train, y_train

# 전체 테스트 데이터 로드
def load_test_data():
    print("테스트 데이터 로드 중...")
    images = []
    labels = []
    skipped = 0

    for i in range(CONFIG['test_num']):
        try:
            imagename = validation_sheet.cell(i+1, 1).value
            label_index = int(validation_sheet.cell(i+1, 2).value)

            if imagename is None or label_index < 0 or label_index >= CONFIG['num_classes']:
                skipped += 1
                continue

            image_path = os.path.join(IMAGESAVEURL_validation, imagename)
            if not os.path.exists(image_path):
                skipped += 1
                continue

            # 이미지 로드
            image = Image.open(image_path)

            # 이미지 크기 확인 및 조정
            width, height = image.size
            if width != 64 or height != 64:
                image = image.resize((64, 64))

            # 이미지 전처리
            imagedata = np.array(image.convert('L'), dtype='float32') / 255.0
            images.append(imagedata)

            # 원-핫 인코딩
            one_hot = np.zeros(CONFIG['num_classes'], dtype=np.float32)
            one_hot[label_index] = 1.0
            labels.append(one_hot)

            if (i+1) % 5000 == 0:
                print(f"처리 중... {i+1}/{CONFIG['test_num']} ({(i+1)/CONFIG['test_num']*100:.1f}%)")

        except Exception as e:
            skipped += 1
            print(f"오류 발생: {e}")
            continue

    print(f"수집된 테스트 데이터: {len(images)}, 스킵됨: {skipped}")

    # 배열로 변환
    x_test = np.array(images)
    y_test = np.array(labels)

    # 채널 차원 추가
    x_test = np.expand_dims(x_test, -1)

    return x_test, y_test

# 모델 정의 (원래대로 유지)
def create_cnn_model(num_classes=CONFIG['num_classes']):
    # 가중치 정규화 추가 (L2 정규화)
    regularizer = tf.keras.regularizers.l2(1e-5)

    # 함수형 API 사용
    inputs = tf.keras.Input(shape=(64, 64, 1))

    # 첫 번째 합성곱 레이어 및 풀링
    x = tf.keras.layers.Conv2D(32, (5, 5), padding='same', activation='relu',
                               kernel_regularizer=regularizer)(inputs)
    x = tf.keras.layers.BatchNormalization()(x)
    x = tf.keras.layers.MaxPooling2D((2, 2))(x)

    # 두 번째 합성곱 레이어 및 풀링
    x = tf.keras.layers.Conv2D(64, (3, 3), padding='same', activation='relu',
                              kernel_regularizer=regularizer)(x)
    x = tf.keras.layers.BatchNormalization()(x)
    x = tf.keras.layers.MaxPooling2D((2, 2))(x)

    # 세 번째 합성곱 레이어 및 풀링
    x = tf.keras.layers.Conv2D(128, (3, 3), padding='same', activation='relu',
                              kernel_regularizer=regularizer)(x)
    x = tf.keras.layers.BatchNormalization()(x)
    x = tf.keras.layers.MaxPooling2D((2, 2))(x)

    # 네 번째 합성곱 레이어 및 풀링
    x = tf.keras.layers.Conv2D(256, (3, 3), padding='same', activation='relu',
                              kernel_regularizer=regularizer)(x)
    x = tf.keras.layers.BatchNormalization()(x)
    x = tf.keras.layers.MaxPooling2D((2, 2))(x)

    # 플래튼 레이어
    x = tf.keras.layers.Flatten()(x)

    # 완전 연결 레이어 1
    x = tf.keras.layers.Dense(2560, activation='relu',
                             kernel_regularizer=regularizer)(x)
    x = tf.keras.layers.BatchNormalization()(x)
    x = tf.keras.layers.Dropout(0.6)(x)  # 논문 그대로 유지

    # 완전 연결 레이어 2 (출력 레이어)
    outputs = tf.keras.layers.Dense(num_classes, activation='softmax')(x)

    # 모델 생성
    model = tf.keras.Model(inputs=inputs, outputs=outputs)

    return model

# 학습률 스케줄러 (원래대로 유지)
def warmup_exp_decay(epoch, lr):
    warmup_epochs = 2
    initial_learning_rate = 1e-6  # 논문 그대로
    if epoch < warmup_epochs:
        # 처음 2 에폭은 학습률 워밍업
        return initial_learning_rate * 5 * ((warmup_epochs - epoch) / warmup_epochs + 1)
    else:
        # 나머지 에폭에서는 지수 감소
        return initial_learning_rate * (0.95 ** (epoch - warmup_epochs))

# 모델 학습 함수 (청킹 제거)
def train_model(x_train, y_train, x_test, y_test, epochs=15):
    # 모델 생성
    model = create_cnn_model()

    # 모델 컴파일 - 학습률 초기값 설정
    model.compile(
        optimizer=tf.keras.optimizers.Adam(learning_rate=1e-6),
        loss='categorical_crossentropy',
        metrics=['accuracy']
    )

    # 모델 요약
    model.summary()

    # 콜백 정의
    lr_scheduler = tf.keras.callbacks.LearningRateScheduler(warmup_exp_decay)
    early_stop = tf.keras.callbacks.EarlyStopping(
        monitor='val_accuracy',
        patience=3,
        restore_best_weights=True,
        verbose=1
    )
    checkpoint = tf.keras.callbacks.ModelCheckpoint(
        f'best_model_{CONFIG["model_prefix"]}.h5',
        monitor='val_accuracy',
        save_best_only=True,
        verbose=1
    )

    # 모델 훈련
    history = model.fit(
        x_train, y_train,
        batch_size=CONFIG['batch_size'],
        epochs=epochs,
        validation_data=(x_test, y_test),
        callbacks=[lr_scheduler, early_stop, checkpoint],
        verbose=1
    )

    # 최종 모델 평가
    print("\n학습 완료! 최종 모델 평가 중...")
    final_loss, final_acc = model.evaluate(x_test, y_test, verbose=1)
    print(f"최종 테스트 정확도: {final_acc:.4f}")

    return model, history

# 혼동 행렬 생성 및 시각화
def create_confusion_matrix(model, x_test, y_test, name_prefix=CONFIG['model_prefix']):
    print("\n혼동 행렬 생성 중...")

    # 예측 수행
    y_pred = model.predict(x_test)

    # 인덱스로 변환
    y_true = np.argmax(y_test, axis=1)
    y_pred = np.argmax(y_pred, axis=1)

    # 혼동 행렬 계산
    cm = confusion_matrix(y_true, y_pred, labels=range(CONFIG['num_classes']))

    # 클래스별 정확도 계산
    print("\n클래스별 최종 정확도:")
    for c in range(CONFIG['num_classes']):
        mask = y_true == c
        if np.sum(mask) > 0:
            class_acc = np.mean(y_pred[mask] == c)
            print(f"  클래스 {c} ({CONFIG['class_labels'][c]}): {class_acc:.4f} ({np.sum(mask)} 샘플)")

    # 퍼센트로 변환
    cm_sum = cm.sum(axis=1, keepdims=True)
    cm_percentage = np.divide(cm.astype(float), cm_sum, where=(cm_sum != 0)) * 100
    cm_percentage = np.nan_to_num(cm_percentage)

    # 시각화
    plt.figure(figsize=(10, 8))

    sns.heatmap(cm_percentage, annot=True, fmt=".2f", cmap="gray_r",
                xticklabels=CONFIG['class_labels'], yticklabels=CONFIG['class_labels'], cbar=True)

    plt.xlabel("Predict label", fontsize=12)
    plt.ylabel("Actual label", fontsize=12)
    plt.xticks(rotation=45)
    plt.yticks(rotation=45)
    plt.title("Confusion Matrix (%)", fontsize=14)

    plt.tight_layout()
    plt.savefig(f'{name_prefix}_confusion_matrix.png')
    print(f"혼동 행렬이 '{name_prefix}_confusion_matrix.png'로 저장되었습니다.")

    # 메모리 정리
    plt.close()

    # 모델 저장
    model.save(f'{name_prefix}_cnn_model.h5')
    print(f"모델이 {name_prefix}_cnn_model.h5로 저장되었습니다.")

# 학습 곡선 시각화
def plot_learning_curves(history, name_prefix=CONFIG['model_prefix']):
    plt.figure(figsize=(12, 5))

    # 정확도 곡선
    plt.subplot(1, 2, 1)
    plt.plot(history.history['accuracy'], label='Train')
    plt.plot(history.history['val_accuracy'], label='Validation')
    plt.title('Model Accuracy')
    plt.xlabel('Epoch')
    plt.ylabel('Accuracy')
    plt.legend()

    # 손실 곡선
    plt.subplot(1, 2, 2)
    plt.plot(history.history['loss'], label='Train')
    plt.plot(history.history['val_loss'], label='Validation')
    plt.title('Model Loss')
    plt.xlabel('Epoch')
    plt.ylabel('Loss')
    plt.legend()

    plt.tight_layout()
    plt.savefig(f'{name_prefix}_learning_curves.png')
    print(f"학습 곡선이 '{name_prefix}_learning_curves.png'로 저장되었습니다.")

    # 메모리 정리
    plt.close()

# 메인 실행 함수
def main():
    print("전체 데이터 로드 및 학습 시작...")
    start_time = time.time()

    # 데이터 로드
    x_train, y_train = load_training_data()
    x_test, y_test = load_test_data()

    print(f"훈련 데이터 크기: {x_train.shape}, 테스트 데이터 크기: {x_test.shape}")

    # 모델 학습
    model, history = train_model(x_train, y_train, x_test, y_test, epochs=15)

    # 학습 곡선 시각화
    plot_learning_curves(history)

    # 혼동 행렬 생성 및 시각화
    create_confusion_matrix(model, x_test, y_test)

    total_time = time.time() - start_time
    print(f"\n실행이 모두 완료되었습니다. 총 소요 시간: {total_time:.2f}초 ({total_time/60:.2f}분)")

if __name__ == "__main__":
    main()
