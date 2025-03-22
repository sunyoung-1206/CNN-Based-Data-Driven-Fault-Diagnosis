# 문제점 해결: 함수 매개변수 문제 수정

import tensorflow as tf
import numpy as np
import pandas as pd
import os
from PIL import Image
import openpyxl
import random
import time
import matplotlib.pyplot as plt
import seaborn as sns
from sklearn.metrics import confusion_matrix

# 로깅 레벨 설정
os.environ['TF_CPP_MIN_LOG_LEVEL'] = '2'

# 경로 설정
IMAGESAVEURL_training = "/home/potato/박선영/CNN-Based-Data-Driven-Fault-Diagnosis/generated_data(2000)/train_images"
IMAGESAVEURL_validation = "/home/potato/박선영/CNN-Based-Data-Driven-Fault-Diagnosis/generated_data(2000)/test_images"

# Excel 파일 로드
training_wb = openpyxl.load_workbook('/home/potato/박선영/CNN-Based-Data-Driven-Fault-Diagnosis/generated_data(2000)/trainingImageList.xlsx')
training_sheet = training_wb.active
validation_wb = openpyxl.load_workbook('/home/potato/박선영/CNN-Based-Data-Driven-Fault-Diagnosis/generated_data(2000)/validationImageList.xlsx')
validation_sheet = validation_wb.active

# 데이터 배열 초기화
train_num = 80000
test_num = 16000

# 데이터 로드 및 전처리 함수 - 매개변수 수정
def load_data(sheet, img_dir, num_samples):
    images = []
    labels = []
    skipped = 0
    processed = 0

    print(f"데이터 처리 시작 ({img_dir})...")
    start_time = time.time()

    # 전체 데이터를 메모리에 로드
    data_entries = []
    for i in range(num_samples):
        try:
            imagename = sheet.cell(i+1, 1).value
            label_index = int(sheet.cell(i+1, 2).value)
            if imagename is None or label_index < 0 or label_index >= 10:
                skipped += 1
                continue

            image_path = os.path.join(img_dir, imagename)
            if not os.path.exists(image_path):
                skipped += 1
                continue

            data_entries.append((image_path, label_index))
            processed += 1

            if (i+1) % 5000 == 0:
                print(f"데이터 처리 중... {i+1}/{num_samples} ({(i+1)/num_samples*100:.1f}%)")

        except Exception as e:
            skipped += 1
            continue

    # 데이터 셔플
    random.shuffle(data_entries)

    # 셔플된 데이터 로드
    for img_path, label in data_entries:
        try:
            image = Image.open(img_path)
            # 이미지 크기 확인 및 조정
            width, height = image.size
            if width != 64 or height != 64:
                image = image.resize((64, 64))

            # 이미지 전처리
            imagedata = np.array(image.convert('L'), dtype='float32') / 255.0

            # 결과 저장
            images.append(imagedata)

            # 원-핫 인코딩
            one_hot = np.zeros(10, dtype=np.float32)
            one_hot[label] = 1.0
            labels.append(one_hot)

        except Exception as e:
            # 이미 유효성 검사를 통과한 데이터이므로 여기서 오류가 발생하면 문제가 있음
            print(f"이미지 로드 중 오류 발생: {img_path}, {str(e)}")

    # 배열로 변환
    images = np.array(images)
    labels = np.array(labels)

    print(f"데이터 처리 완료 ({time.time() - start_time:.2f}초)")
    print(f"처리된 이미지: {processed}/{num_samples}, 스킵된 이미지: {skipped}/{num_samples}")

    return images, labels

# 데이터 통계 확인 함수
def check_data_distribution(y_data):
    label_counts = np.sum(y_data, axis=0)
    total_samples = np.sum(label_counts)

    print("\n데이터 분포 확인:")
    print("=" * 50)
    for i, count in enumerate(label_counts):
        percentage = (count / total_samples) * 100
        print(f"클래스 {i}: {int(count)} 샘플 ({percentage:.2f}%)")
    print("=" * 50)

    # 클래스 불균형 확인
    if max(label_counts) / (min(label_counts) + 1e-10) > 5:
        print("⚠ 경고: 심각한 클래스 불균형이 감지되었습니다.")
    return label_counts

def validate_images(x_data, sample_size=5):
    empty_images = np.sum(np.all(x_data == 0, axis=(1, 2)))
    if empty_images > 0:
        print(f"⚠ 경고: {empty_images}개의 빈 이미지가 발견되었습니다.")

    min_val = np.min(x_data)
    max_val = np.max(x_data)
    mean_val = np.mean(x_data)
    std_val = np.std(x_data)

    print("\n이미지 통계:")
    print("=" * 50)
    print(f"최소값: {min_val}, 최대값: {max_val}")
    print(f"평균값: {mean_val:.4f}, 표준편차: {std_val:.4f}")

    if min_val < 0 or max_val > 1:
        print("⚠ 경고: 이미지 픽셀 값이 [0, 1] 범위를 벗어났습니다.")

    # 샘플 이미지의 평균과 표준편차
    indices = random.sample(range(len(x_data)), min(sample_size, len(x_data)))
    for i, idx in enumerate(indices):
        img = x_data[idx]
        print(f"샘플 {i+1}: 평균={np.mean(img):.4f}, 표준편차={np.std(img):.4f}")
    print("=" * 50)

# 데이터 로드 (셔플 적용) - 매개변수 수정
x_data, y_data = load_data(training_sheet, IMAGESAVEURL_training, train_num)
x_test, y_test = load_data(validation_sheet, IMAGESAVEURL_validation, test_num)

# 데이터 검증
train_label_counts = check_data_distribution(y_data)
validate_images(x_data)
test_label_counts = check_data_distribution(y_test)
validate_images(x_test)

# CNN 모델 정의 함수
def create_cnn_model():
    # 가중치 정규화 추가 (L2 정규화)
    regularizer = tf.keras.regularizers.l2(1e-5)

    # 함수형 API 사용
    inputs = tf.keras.Input(shape=(64, 64, 1))

    # 첫 번째 합성곱 레이어 및 풀링
    x = tf.keras.layers.Conv2D(32, (5, 5), padding='same', activation='relu',
                               kernel_regularizer=regularizer)(inputs)
    x = tf.keras.layers.BatchNormalization()(x)
    x = tf.keras.layers.MaxPooling2D((2, 2))(x)

    # 두 번째 합성곱 레이어 및 풀링
    x = tf.keras.layers.Conv2D(64, (3, 3), padding='same', activation='relu',
                              kernel_regularizer=regularizer)(x)
    x = tf.keras.layers.BatchNormalization()(x)
    x = tf.keras.layers.MaxPooling2D((2, 2))(x)

    # 세 번째 합성곱 레이어 및 풀링
    x = tf.keras.layers.Conv2D(128, (3, 3), padding='same', activation='relu',
                              kernel_regularizer=regularizer)(x)
    x = tf.keras.layers.BatchNormalization()(x)
    x = tf.keras.layers.MaxPooling2D((2, 2))(x)

    # 네 번째 합성곱 레이어 및 풀링
    x = tf.keras.layers.Conv2D(256, (3, 3), padding='same', activation='relu',
                              kernel_regularizer=regularizer)(x)
    x = tf.keras.layers.BatchNormalization()(x)
    x = tf.keras.layers.MaxPooling2D((2, 2))(x)

    # 플래튼 레이어
    x = tf.keras.layers.Flatten()(x)

    # 완전 연결 레이어 1
    x = tf.keras.layers.Dense(2560, activation='relu',
                             kernel_regularizer=regularizer)(x)
    x = tf.keras.layers.BatchNormalization()(x)
    x = tf.keras.layers.Dropout(0.6)(x)  # 논문 그대로 유지

    # 완전 연결 레이어 2 (출력 레이어)
    outputs = tf.keras.layers.Dense(10, activation='softmax')(x)

    # 모델 생성
    model = tf.keras.Model(inputs=inputs, outputs=outputs)

    return model

# 모델 생성
model = create_cnn_model()

# 학습률 워밍업 + 지수 감소 스케줄
def warmup_exp_decay(epoch, lr):
    warmup_epochs = 2
    initial_learning_rate = 1e-6  # 논문 그대로
    if epoch < warmup_epochs:
        # 처음 2 에폭은 학습률 워밍업
        return initial_learning_rate * 5 * ((warmup_epochs - epoch) / warmup_epochs + 1)
    else:
        # 나머지 에폭에서는 지수 감소
        return initial_learning_rate * (0.95 ** (epoch - warmup_epochs))

# 모델 컴파일
model.compile(
    optimizer=tf.keras.optimizers.Adam(learning_rate=1e-6),  # 초기값 설정
    loss='categorical_crossentropy',
    metrics=['accuracy']
)
model.summary()

# 데이터 증강 설정
data_augmentation = tf.keras.Sequential([
    tf.keras.layers.RandomRotation(0.05),
    tf.keras.layers.RandomZoom(0.05),
    tf.keras.layers.RandomTranslation(0.05, 0.05),
    tf.keras.layers.RandomFlip("horizontal")
])

# 데이터 증강 함수
@tf.function
def augment_data(image, label):
    # 50% 확률로 증강 적용
    if tf.random.uniform(shape=()) < 0.5:
        image = data_augmentation(image, training=True)
    return image, label

# 학습용 데이터셋 생성
batch_size = 32
train_dataset = tf.data.Dataset.from_tensor_slices((
    np.expand_dims(x_data, -1),  # 채널 차원 추가
    y_data
))
train_dataset = train_dataset.shuffle(len(x_data), reshuffle_each_iteration=True)
train_dataset = train_dataset.map(augment_data, num_parallel_calls=tf.data.AUTOTUNE)
train_dataset = train_dataset.batch(batch_size).prefetch(tf.data.AUTOTUNE)

# 테스트용 데이터셋 생성
test_dataset = tf.data.Dataset.from_tensor_slices((
    np.expand_dims(x_test, -1),
    y_test
)).batch(batch_size)

# 학습률 스케줄러 콜백
lr_scheduler = tf.keras.callbacks.LearningRateScheduler(warmup_exp_decay)

# 모델 체크포인트 설정
checkpoint = tf.keras.callbacks.ModelCheckpoint(
    'best_model.h5',
    monitor='val_accuracy',
    save_best_only=True,
    mode='max',
    verbose=1
)

# 학습률 모니터링 콜백 (수정됨)
class LrMonitorCallback(tf.keras.callbacks.Callback):
    def on_epoch_end(self, epoch, logs=None):
        # TF 2.x에서 올바른 학습률 접근 방식
        try:
            # TensorFlow 2.x에서 현재 학습률에 접근하는 방법
            if hasattr(self.model.optimizer, '_decayed_lr'):
                lr = self.model.optimizer._decayed_lr(tf.float32).numpy()
            elif hasattr(self.model.optimizer, 'lr'):
                lr = tf.keras.backend.get_value(self.model.optimizer.lr)
            elif hasattr(self.model.optimizer, 'learning_rate'):
                lr = tf.keras.backend.get_value(self.model.optimizer.learning_rate)
            else:
                lr = 0.0
                print("경고: 학습률을 확인할 수 없습니다.")

            print(f"\n현재 학습률: {lr:.10f}")
        except Exception as e:
            print(f"\n학습률 출력 오류: {str(e)}")

# 검증 콜백
class ValidationCallback(tf.keras.callbacks.Callback):
    def __init__(self, validation_data):
        super(ValidationCallback, self).__init__()
        self.validation_data = validation_data

    def on_epoch_end(self, epoch, logs=None):
        # 검증 데이터에 대한 성능 평가
        val_loss, val_acc = self.model.evaluate(self.validation_data, verbose=0)
        print(f"\n에폭 {epoch+1} 검증 성능 - 정확도: {val_acc:.4f}, 손실: {val_loss:.4f}")

        # 클래스별 정확도 확인 (샘플링)
        all_predictions = []
        all_labels = []

        # 에러 처리 추가
        try:
            # 검증 데이터의 일부만 예측 (너무 많은 배치를 처리하지 않도록)
            for i, (x_batch, y_batch) in enumerate(self.validation_data):
                if i >= 5:  # 최대 5개 배치만 처리
                    break

                # 예측 수행
                preds = self.model.predict(x_batch, verbose=0)
                pred_labels = np.argmax(preds, axis=1)
                true_labels = np.argmax(y_batch.numpy(), axis=1)

                all_predictions.extend(pred_labels)
                all_labels.extend(true_labels)

            # 클래스별 정확도 계산
            print("클래스별 정확도 (샘플링):")
            for c in range(10):
                mask = np.array(all_labels) == c
                if np.sum(mask) > 0:
                    class_acc = np.mean(np.array(all_predictions)[mask] == c)
                    print(f"  클래스 {c}: {class_acc:.4f}")
                else:
                    print(f"  클래스 {c}: 샘플 없음")
        except Exception as e:
            print(f"검증 중 오류 발생: {str(e)}")

# 조기 종료 설정
early_stopping = tf.keras.callbacks.EarlyStopping(
    monitor='val_loss',
    patience=3,
    restore_best_weights=True
)

# 모니터링 콜백 생성
val_callback = ValidationCallback(test_dataset)
lr_monitor = LrMonitorCallback()

# 모델 학습
print("\n모델 학습 시작...")
epochs = 15  # 에폭 수 증가
history = model.fit(
    train_dataset,
    epochs=epochs,
    validation_data=test_dataset,
    callbacks=[checkpoint, lr_scheduler, val_callback, lr_monitor, early_stopping],
    verbose=1
)

# 최적 모델 로드 및 평가
try:
    # 최적 모델 로드
    best_model = tf.keras.models.load_model('best_model.h5')

    # 테스트 데이터로 모델 평가
    print("\n최종 모델 평가...")
    test_loss, test_accuracy = best_model.evaluate(test_dataset)
    print(f"테스트 정확도: {test_accuracy:.4f}")

    # 전체 테스트 데이터에 대한 예측
    print("\n클래스별 정확도 계산 중...")
    predictions = best_model.predict(test_dataset)
    y_pred = np.argmax(predictions, axis=1)

    # 실제 레이블
    all_y_test = []
    for _, y_batch in test_dataset:
        all_y_test.extend(np.argmax(y_batch.numpy(), axis=1))
    y_true = np.array(all_y_test)

    # 클래스별 정확도 계산
    print("\n클래스별 최종 정확도:")
    for c in range(10):
        mask = y_true == c
        if np.sum(mask) > 0:
            class_acc = np.mean(y_pred[mask] == c)
            print(f"  클래스 {c}: {class_acc:.4f} ({np.sum(mask)} 샘플)")

    # 혼동 행렬 생성 및 시각화
    print("\n혼동 행렬 생성 중...")
    # Confusion Matrix 계산
    cm = confusion_matrix(y_true, y_pred)

    # 퍼센트 값으로 변환 (0으로 나누는 문제 방지)
    cm_sum = cm.sum(axis=1, keepdims=True)
    cm_percentage = np.divide(cm.astype(float), cm_sum, where=(cm_sum != 0)) * 100  # 0으로 나누는 것 방지
    cm_percentage = np.nan_to_num(cm_percentage)  # NaN을 0으로 변환

    # 클래스 라벨 정의
    labels = ["BF0.18", "BF0.36", "BF0.54", "OF0.18", "OF0.36", "OF0.54", "IF0.18", "IF0.36", "IF0.54", "NO"]

    plt.figure(figsize=(8, 6))

    # 히트맵 그리기
    sns.heatmap(cm_percentage, annot=True, fmt=".2f", cmap="gray_r",
                xticklabels=labels, yticklabels=labels, cbar=True)

    # 축 레이블 추가
    plt.xlabel("Predict label", fontsize=12)
    plt.ylabel("Actual label", fontsize=12)

    # 회전된 라벨 스타일 적용
    plt.xticks(rotation=45)
    plt.yticks(rotation=45)

    # 그래프 제목 설정
    plt.title("Confusion Matrix (%)", fontsize=14)

    # 그래프 저장
    plt.tight_layout()
    plt.savefig('confusion_matrix.png')
    print("혼동 행렬이 'confusion_matrix.png'로 저장되었습니다.")

    # 그래프 표시
    plt.show()

    # 모델 저장
    best_model.save('cnn_model.h5')
    print("모델이 cnn_model.h5로 저장되었습니다.")
except Exception as e:
    print(f"최적 모델 로드 중 오류 발생: {str(e)}")
    print("현재 모델 사용...")

    # 현재 모델 평가
    test_loss, test_accuracy = model.evaluate(test_dataset)
    print(f"테스트 정확도: {test_accuracy:.4f}")

    # 전체 테스트 데이터에 대한 예측
    print("\n클래스별 정확도 계산 중...")
    predictions = model.predict(test_dataset)
    y_pred = np.argmax(predictions, axis=1)

    # 실제 레이블
    all_y_test = []
    for _, y_batch in test_dataset:
        all_y_test.extend(np.argmax(y_batch.numpy(), axis=1))
    y_true = np.array(all_y_test)

    # 혼동 행렬 생성 및 시각화
    print("\n혼동 행렬 생성 중...")
    # Confusion Matrix 계산
    cm = confusion_matrix(y_true, y_pred)

    # 퍼센트 값으로 변환 (0으로 나누는 문제 방지)
    cm_sum = cm.sum(axis=1, keepdims=True)
    cm_percentage = np.divide(cm.astype(float), cm_sum, where=(cm_sum != 0)) * 100  # 0으로 나누는 것 방지
    cm_percentage = np.nan_to_num(cm_percentage)  # NaN을 0으로 변환

    # 클래스 라벨 정의
    labels = ["BF0.18", "BF0.36", "BF0.54", "OF0.18", "OF0.36", "OF0.54", "IF0.18", "IF0.36", "IF0.54", "NO"]

    plt.figure(figsize=(8, 6))

    # 히트맵 그리기
    sns.heatmap(cm_percentage, annot=True, fmt=".2f", cmap="gray_r",
                xticklabels=labels, yticklabels=labels, cbar=True)

    # 축 레이블 추가
    plt.xlabel("Predict label", fontsize=12)
    plt.ylabel("Actual label", fontsize=12)

    # 회전된 라벨 스타일 적용
    plt.xticks(rotation=45)
    plt.yticks(rotation=45)

    # 그래프 제목 설정
    plt.title("Confusion Matrix (%)", fontsize=14)

    # 그래프 저장
    plt.tight_layout()
    plt.savefig('confusion_matrix.png')
    print("혼동 행렬이 'confusion_matrix.png'로 저장되었습니다.")

    # 그래프 표시
    plt.show()

    # 모델 저장
    model.save('cnn_model.h5')
    print("모델이 cnn_model.h5로 저장되었습니다.")

print("\n실행이 모두 완료되었습니다.")
