#!/usr/bin/env python3
"""
test1_v4_memory_efficient.py - 모델 파라미터는 그대로 유지하면서 메모리 효율성 개선
"""

import tensorflow as tf
import numpy as np
import pandas as pd
import os
from PIL import Image
import openpyxl
import random
import time
import matplotlib.pyplot as plt
import seaborn as sns
from sklearn.metrics import confusion_matrix
import gc

# 메모리 제한 설정 (필요시 활성화)
gpus = tf.config.experimental.list_physical_devices('GPU')
if gpus:
    try:
        for gpu in gpus:
            tf.config.experimental.set_memory_growth(gpu, True)
        print(f"{len(gpus)}개 GPU에 메모리 성장 제한 설정 완료")
    except RuntimeError as e:
        print(f"GPU 설정 오류: {e}")

# 로깅 레벨 설정
os.environ['TF_CPP_MIN_LOG_LEVEL'] = '2'

# 경로 설정
IMAGESAVEURL_training = "/home/potato/박선영/CNN-Based-Data-Driven-Fault-Diagnosis/test1_generated_data/train_images"
IMAGESAVEURL_validation = "/home/potato/박선영/CNN-Based-Data-Driven-Fault-Diagnosis/test1_generated_data/test_images"

# Excel 파일 로드
training_wb = openpyxl.load_workbook('/home/potato/박선영/CNN-Based-Data-Driven-Fault-Diagnosis/test1_generated_data/trainingImageList.xlsx')
training_sheet = training_wb.active
validation_wb = openpyxl.load_workbook('/home/potato/박선영/CNN-Based-Data-Driven-Fault-Diagnosis/test1_generated_data/validationImageList.xlsx')
validation_sheet = validation_wb.active

# 데이터 배열 초기화 - 상단에 설정하여 쉽게 변경 가능
CONFIG = {
    'train_num': 80000,     # 원래 값 유지
    'test_num': 16000,      # 원래 값 유지
    'num_classes': 10,      # 분류할 클래스 수
    'model_prefix': 'test1_v4',  # 저장 파일 접두사
    'class_labels': ["BF0.18", "BF0.36", "BF0.54", "OF0.18", "OF0.36", "OF0.54", "IF0.18", "IF0.36", "IF0.54", "NO"],  # 혼동 행렬에 표시할 클래스 라벨
    'batch_size': 32,       # 원래 배치 크기 유지
    'chunks': 4             # 데이터를 나눌 청크 수 (메모리 사용량에 따라 조정)
}

# 청크 기반 데이터 로드 (훈련 데이터)
def load_training_data_by_chunks():
    # 전체 훈련 데이터 경로와 라벨 수집
    print("훈련 데이터 경로 및 라벨 수집 중...")
    data_entries = []
    skipped = 0

    for i in range(CONFIG['train_num']):
        try:
            imagename = training_sheet.cell(i+1, 1).value
            label_index = int(training_sheet.cell(i+1, 2).value)

            if imagename is None or label_index < 0 or label_index >= CONFIG['num_classes']:
                skipped += 1
                continue

            image_path = os.path.join(IMAGESAVEURL_training, imagename)
            if not os.path.exists(image_path):
                skipped += 1
                continue

            data_entries.append((image_path, label_index))

            if (i+1) % 10000 == 0:
                print(f"처리 중... {i+1}/{CONFIG['train_num']} ({(i+1)/CONFIG['train_num']*100:.1f}%)")

        except Exception as e:
            skipped += 1
            continue

    # 데이터 셔플
    random.shuffle(data_entries)

    print(f"수집된 훈련 데이터: {len(data_entries)}, 스킵됨: {skipped}")

    # 청크로 분할
    chunk_size = len(data_entries) // CONFIG['chunks']
    chunks = []

    for i in range(CONFIG['chunks']):
        start_idx = i * chunk_size
        end_idx = start_idx + chunk_size if i < CONFIG['chunks'] - 1 else len(data_entries)
        chunks.append(data_entries[start_idx:end_idx])

    print(f"데이터를 {CONFIG['chunks']}개 청크로 분할 완료 (청크당 약 {chunk_size}개 샘플)")

    return chunks

# 청크 기반 데이터 로드 (테스트 데이터)
def load_test_data_by_chunks():
    # 전체 테스트 데이터 경로와 라벨 수집
    print("테스트 데이터 경로 및 라벨 수집 중...")
    data_entries = []
    skipped = 0

    for i in range(CONFIG['test_num']):
        try:
            imagename = validation_sheet.cell(i+1, 1).value
            label_index = int(validation_sheet.cell(i+1, 2).value)

            if imagename is None or label_index < 0 or label_index >= CONFIG['num_classes']:
                skipped += 1
                continue

            image_path = os.path.join(IMAGESAVEURL_validation, imagename)
            if not os.path.exists(image_path):
                skipped += 1
                continue

            data_entries.append((image_path, label_index))

            if (i+1) % 5000 == 0:
                print(f"처리 중... {i+1}/{CONFIG['test_num']} ({(i+1)/CONFIG['test_num']*100:.1f}%)")

        except Exception as e:
            skipped += 1
            continue

    print(f"수집된 테스트 데이터: {len(data_entries)}, 스킵됨: {skipped}")

    # 청크 크기 계산 (테스트 데이터는 더 작은 청크로 분할)
    chunk_size = len(data_entries) // max(2, CONFIG['chunks'] // 2)
    chunks = []

    for i in range(max(2, CONFIG['chunks'] // 2)):
        start_idx = i * chunk_size
        end_idx = start_idx + chunk_size if i < max(2, CONFIG['chunks'] // 2) - 1 else len(data_entries)
        chunks.append(data_entries[start_idx:end_idx])

    print(f"데이터를 {len(chunks)}개 청크로 분할 완료 (청크당 약 {chunk_size}개 샘플)")

    return chunks

# 청크에서 데이터셋 생성
def create_dataset_from_chunk(chunk, is_training=True):
    print(f"청크 처리 중 (샘플 수: {len(chunk)})...")
    start_time = time.time()

    images = []
    labels = []

    for img_path, label in chunk:
        try:
            image = Image.open(img_path)
            # 이미지 크기 확인 및 조정
            width, height = image.size
            if width != 64 or height != 64:
                image = image.resize((64, 64))

            # 이미지 전처리
            imagedata = np.array(image.convert('L'), dtype='float32') / 255.0

            # 결과 저장
            images.append(imagedata)

            # 원-핫 인코딩
            one_hot = np.zeros(CONFIG['num_classes'], dtype=np.float32)
            one_hot[label] = 1.0
            labels.append(one_hot)

        except Exception as e:
            print(f"이미지 로드 중 오류 발생: {img_path}, {str(e)}")

    # 배열로 변환
    x_data = np.array(images)
    y_data = np.array(labels)

    # 메모리 해제
    del images
    del labels
    gc.collect()

    # 데이터셋 생성
    dataset = tf.data.Dataset.from_tensor_slices((
        np.expand_dims(x_data, -1),  # 채널 차원 추가
        y_data
    ))

    if is_training:
        dataset = dataset.shuffle(len(x_data), reshuffle_each_iteration=True)

    dataset = dataset.batch(CONFIG['batch_size']).prefetch(tf.data.AUTOTUNE)

    print(f"청크 처리 완료 ({time.time() - start_time:.2f}초)")

    # 메모리 해제
    del x_data
    del y_data
    gc.collect()

    return dataset

# 모델 정의 (원래대로 유지)
def create_cnn_model(num_classes=CONFIG['num_classes']):
    # 가중치 정규화 추가 (L2 정규화)
    regularizer = tf.keras.regularizers.l2(1e-5)

    # 함수형 API 사용
    inputs = tf.keras.Input(shape=(64, 64, 1))

    # 첫 번째 합성곱 레이어 및 풀링
    x = tf.keras.layers.Conv2D(32, (5, 5), padding='same', activation='relu',
                               kernel_regularizer=regularizer)(inputs)
    x = tf.keras.layers.BatchNormalization()(x)
    x = tf.keras.layers.MaxPooling2D((2, 2))(x)

    # 두 번째 합성곱 레이어 및 풀링
    x = tf.keras.layers.Conv2D(64, (3, 3), padding='same', activation='relu',
                              kernel_regularizer=regularizer)(x)
    x = tf.keras.layers.BatchNormalization()(x)
    x = tf.keras.layers.MaxPooling2D((2, 2))(x)

    # 세 번째 합성곱 레이어 및 풀링
    x = tf.keras.layers.Conv2D(128, (3, 3), padding='same', activation='relu',
                              kernel_regularizer=regularizer)(x)
    x = tf.keras.layers.BatchNormalization()(x)
    x = tf.keras.layers.MaxPooling2D((2, 2))(x)

    # 네 번째 합성곱 레이어 및 풀링
    x = tf.keras.layers.Conv2D(256, (3, 3), padding='same', activation='relu',
                              kernel_regularizer=regularizer)(x)
    x = tf.keras.layers.BatchNormalization()(x)
    x = tf.keras.layers.MaxPooling2D((2, 2))(x)

    # 플래튼 레이어
    x = tf.keras.layers.Flatten()(x)

    # 완전 연결 레이어 1
    x = tf.keras.layers.Dense(2560, activation='relu',
                             kernel_regularizer=regularizer)(x)
    x = tf.keras.layers.BatchNormalization()(x)
    x = tf.keras.layers.Dropout(0.6)(x)  # 논문 그대로 유지

    # 완전 연결 레이어 2 (출력 레이어)
    outputs = tf.keras.layers.Dense(num_classes, activation='softmax')(x)

    # 모델 생성
    model = tf.keras.Model(inputs=inputs, outputs=outputs)

    return model

# 학습률 스케줄러 (원래대로 유지)
def warmup_exp_decay(epoch, lr):
    warmup_epochs = 2
    initial_learning_rate = 1e-6  # 논문 그대로
    if epoch < warmup_epochs:
        # 처음 2 에폭은 학습률 워밍업
        return initial_learning_rate * 5 * ((warmup_epochs - epoch) / warmup_epochs + 1)
    else:
        # 나머지 에폭에서는 지수 감소
        return initial_learning_rate * (0.95 ** (epoch - warmup_epochs))

# 메모리 효율적인 학습 함수
def train_model_by_chunks(train_chunks, test_chunks, epochs=15):
    # 모델 생성
    model = create_cnn_model()

    # 모델 컴파일 - 학습률 초기값 설정
    model.compile(
        optimizer=tf.keras.optimizers.Adam(learning_rate=1e-6),
        loss='categorical_crossentropy',
        metrics=['accuracy']
    )

    # 모델 요약
    model.summary()

    # 테스트 데이터셋 생성 (전체 테스트 데이터는 더 작아서 한 번에 처리 가능)
    print("테스트 데이터셋 생성 중...")
    all_test_entries = []
    for chunk in test_chunks:
        all_test_entries.extend(chunk)

    test_dataset = create_dataset_from_chunk(all_test_entries, is_training=False)

    # 에폭별 학습
    best_val_acc = 0
    patience_counter = 0

    for epoch in range(epochs):
        print(f"\n=== 에폭 {epoch+1}/{epochs} ===")
        epoch_start_time = time.time()

        # 현재 학습률 계산 (학습률 스케줄러 함수 이용)
        current_lr = warmup_exp_decay(epoch, 1e-6)
        print(f"현재 에폭 학습률: {current_lr:.10f}")

        # 옵티마이저 재생성 (학습률 직접 설정)
        # TensorFlow 버전 호환성 문제를 방지하기 위한 방법
        model.optimizer = tf.keras.optimizers.Adam(learning_rate=current_lr)

        # 에폭 내 각 청크에 대해 학습
        train_loss = []
        train_acc = []

        for chunk_idx, chunk in enumerate(train_chunks):
            print(f"\n청크 {chunk_idx+1}/{len(train_chunks)} 학습 중...")

            # 현재 청크에서 데이터셋 생성
            train_dataset = create_dataset_from_chunk(chunk, is_training=True)

            # 현재 청크에 대한 학습
            history = model.fit(
                train_dataset,
                epochs=1,  # 각 청크에 대해 1 에폭씩 학습
                verbose=1
            )

            # 메트릭 수집
            train_loss.append(history.history['loss'][0])
            train_acc.append(history.history['accuracy'][0])

            # 메모리 정리
            del train_dataset
            gc.collect()
            tf.keras.backend.clear_session()

        # 에폭 평균 메트릭 계산
        avg_train_loss = np.mean(train_loss)
        avg_train_acc = np.mean(train_acc)

        # 검증 데이터에 대한 평가
        val_loss, val_acc = model.evaluate(test_dataset, verbose=1)

        # 체크포인트 로직
        if val_acc > best_val_acc:
            best_val_acc = val_acc
            print(f"검증 정확도 향상: {best_val_acc:.4f}, 모델 저장 중...")
            model.save(f'best_model_{CONFIG["model_prefix"]}.h5')
            patience_counter = 0  # 개선되면 카운터 리셋
        else:
            patience_counter += 1  # 개선되지 않으면 카운터 증가

        # 조기 종료 확인
        if patience_counter >= 3:
            print(f"검증 손실이 3번 연속 개선되지 않아 훈련 조기 종료")
            break

        # 에폭 요약
        print(f"\n에폭 {epoch+1} 요약:")
        print(f"훈련 손실: {avg_train_loss:.4f}, 훈련 정확도: {avg_train_acc:.4f}")
        print(f"검증 손실: {val_loss:.4f}, 검증 정확도: {val_acc:.4f}")
        print(f"에폭 소요 시간: {time.time() - epoch_start_time:.2f}초")

        # 메모리 정리
        gc.collect()

    # 최종 모델 평가
    print("\n학습 완료! 최종 모델 평가 중...")
    final_loss, final_acc = model.evaluate(test_dataset, verbose=1)
    print(f"최종 테스트 정확도: {final_acc:.4f}")

    # 최적 모델 로드
    try:
        best_model = tf.keras.models.load_model(f'best_model_{CONFIG["model_prefix"]}.h5')
        print("최적 모델 로드 완료!")
        return best_model
    except Exception as e:
        print(f"최적 모델 로드 오류: {e}")
        return model

# 효율적인 혼동 행렬 생성 및 시각화
def create_confusion_matrix(model, test_chunks, name_prefix=CONFIG['model_prefix']):
    print("\n혼동 행렬 생성 중...")

    # 모든 예측과 실제 레이블 저장
    all_predictions = []
    all_true_labels = []

    # 각 청크에 대해 예측 수행
    for chunk_idx, chunk in enumerate(test_chunks):
        print(f"청크 {chunk_idx+1}/{len(test_chunks)} 처리 중...")

        # 현재 청크에서 데이터셋 생성
        test_dataset = create_dataset_from_chunk(chunk, is_training=False)

        # 예측 수행
        for x_batch, y_batch in test_dataset:
            # 배치 예측
            y_pred_batch = model.predict(x_batch, verbose=0)

            # 인덱스로 변환
            y_true_batch = tf.argmax(y_batch, axis=1).numpy()
            y_pred_batch = tf.argmax(y_pred_batch, axis=1).numpy()

            # 결과 저장
            all_predictions.extend(y_pred_batch)
            all_true_labels.extend(y_true_batch)

        # 메모리 정리
        del test_dataset
        gc.collect()

    # 혼동 행렬 계산
    cm = confusion_matrix(all_true_labels, all_predictions, labels=range(CONFIG['num_classes']))

    # 클래스별 정확도 계산
    print("\n클래스별 최종 정확도:")
    for c in range(CONFIG['num_classes']):
        mask = np.array(all_true_labels) == c
        if np.sum(mask) > 0:
            class_acc = np.mean(np.array(all_predictions)[mask] == c)
            print(f"  클래스 {c} ({CONFIG['class_labels'][c]}): {class_acc:.4f} ({np.sum(mask)} 샘플)")

    # 퍼센트로 변환
    cm_sum = cm.sum(axis=1, keepdims=True)
    cm_percentage = np.divide(cm.astype(float), cm_sum, where=(cm_sum != 0)) * 100
    cm_percentage = np.nan_to_num(cm_percentage)

    # 시각화
    plt.figure(figsize=(10, 8))

    sns.heatmap(cm_percentage, annot=True, fmt=".2f", cmap="gray_r",
                xticklabels=CONFIG['class_labels'], yticklabels=CONFIG['class_labels'], cbar=True)

    plt.xlabel("Predict label", fontsize=12)
    plt.ylabel("Actual label", fontsize=12)
    plt.xticks(rotation=45)
    plt.yticks(rotation=45)
    plt.title("Confusion Matrix (%)", fontsize=14)

    plt.tight_layout()
    plt.savefig(f'{name_prefix}_confusion_matrix.png')
    print(f"혼동 행렬이 '{name_prefix}_confusion_matrix.png'로 저장되었습니다.")

    # 메모리 정리
    plt.close()

    # 모델 저장
    model.save(f'{name_prefix}_cnn_model.h5')
    print(f"모델이 {name_prefix}_cnn_model.h5로 저장되었습니다.")

# 메인 실행 함수
def main():
    print("메모리 효율적인 방식으로 학습 시작...")

    # 데이터 청크 로드
    train_chunks = load_training_data_by_chunks()
    test_chunks = load_test_data_by_chunks()

    # 모델 학습
    model = train_model_by_chunks(train_chunks, test_chunks, epochs=15)

    # 혼동 행렬 생성 및 시각화
    create_confusion_matrix(model, test_chunks)

    print("\n실행이 모두 완료되었습니다.")

if __name__ == "__main__":
    main()
