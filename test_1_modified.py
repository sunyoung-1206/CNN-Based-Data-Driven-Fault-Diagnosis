# 0. 라이브러리 설치
# pip install tensorflow
# pip install numpy
# pip install pandas
# pip install openpyxl

# 1. 라이브러리 임포트
# -*- coding: utf-8 -*-
"""
Original created on Tue May 1 15:31:04 2018
@author: zhaoyuzhi
Updated to modern TensorFlow API in 2025
"""

import tensorflow as tf
import numpy as np
import pandas as pd
import os
from PIL import Image
import openpyxl  # xlrd 대신 openpyxl 사용
import random  # 데이터 셔플을 위해 추가

# 로깅 레벨 설정
os.environ['TF_CPP_MIN_LOG_LEVEL'] = '2'

# GPU 메모리 설정 (OOM 방지)
gpus = tf.config.experimental.list_physical_devices('GPU')
if gpus:
    try:
        for gpu in gpus:
            tf.config.experimental.set_memory_growth(gpu, True)
    except RuntimeError as e:
        print(e)

# 2. 경로 설정
# 경로 설정 (필요에 따라 수정)
IMAGESAVEURL_training = "/home/potato/박선영/CNN-Based-Data-Driven-Fault-Diagnosis/generated_data(2000)/train_images"
IMAGESAVEURL_validation = "/home/potato/박선영/CNN-Based-Data-Driven-Fault-Diagnosis/generated_data(2000)/test_images"

# 3. Excel 파일 로드
# Excel 파일 로드 (xlrd 대신 openpyxl 사용)
training_wb = openpyxl.load_workbook('/home/potato/박선영/CNN-Based-Data-Driven-Fault-Diagnosis/generated_data(2000)/trainingImageList.xlsx')
training_sheet = training_wb.active
validation_wb = openpyxl.load_workbook('/home/potato/박선영/CNN-Based-Data-Driven-Fault-Diagnosis/generated_data(2000)/validationImageList.xlsx')
validation_sheet = validation_wb.active

# 4. 데이터 배열 초기화
# 데이터 저장용 배열 초기화
train_num = 80000
test_num = 16000

# 데이터 제너레이터 함수 정의 (메모리 효율적인 처리를 위해)
def image_generator(sheet, image_url, batch_size, is_training=True):
    """이미지와 레이블을 배치 단위로 생성하는 제너레이터"""
    num_samples = train_num if is_training else test_num
    indices = list(range(1, num_samples + 1))

    if is_training:
        random.shuffle(indices)

    num_batches = (num_samples + batch_size - 1) // batch_size  # 올림 나눗셈

    for batch_idx in range(num_batches):
        start_idx = batch_idx * batch_size
        end_idx = min(start_idx + batch_size, num_samples)
        current_indices = indices[start_idx:end_idx]

        batch_x = np.zeros((len(current_indices), 64, 64, 1), dtype=np.float32)
        batch_y = np.zeros((len(current_indices), 10), dtype=np.float32)

        for i, idx in enumerate(current_indices):
            try:
                imagename = sheet.cell(idx, 1).value
                if imagename is None:
                    continue

                label_index = int(sheet.cell(idx, 2).value)

                # 이미지 로드 및 전처리
                image = Image.open(os.path.join(image_url, imagename))
                imagedata = np.array(image.convert('L'), dtype='float32') / 255.0

                # 데이터 저장
                batch_x[i, :, :, 0] = imagedata
                batch_y[i, label_index] = 1

            except Exception as e:
                print(f"이미지 로드 오류 (행 {idx}): {e}")

        yield batch_x, batch_y

# 7. CNN 모델 정의 함수
# 모델 정의
def create_cnn_model():
    # 함수형 API 사용
    inputs = tf.keras.Input(shape=(64, 64, 1))

    # 첫 번째 합성곱 레이어 및 풀링
    x = tf.keras.layers.Conv2D(32, (5, 5), padding='same', activation='relu')(inputs)
    x = tf.keras.layers.MaxPooling2D((2, 2))(x)

    # 두 번째 합성곱 레이어 및 풀링
    x = tf.keras.layers.Conv2D(64, (3, 3), padding='same', activation='relu')(x)
    x = tf.keras.layers.MaxPooling2D((2, 2))(x)

    # 세 번째 합성곱 레이어 및 풀링
    x = tf.keras.layers.Conv2D(128, (3, 3), padding='same', activation='relu')(x)
    x = tf.keras.layers.MaxPooling2D((2, 2))(x)

    # 네 번째 합성곱 레이어 및 풀링
    x = tf.keras.layers.Conv2D(256, (3, 3), padding='same', activation='relu')(x)
    x = tf.keras.layers.MaxPooling2D((2, 2))(x)

    # 플래튼 레이어
    x = tf.keras.layers.Flatten()(x)

    # 완전 연결 레이어 1
    x = tf.keras.layers.Dense(2560, activation='relu')(x)
    x = tf.keras.layers.Dropout(0.6)(x)  # keep_prob 0.4 -> dropout 0.6

    # 완전 연결 레이어 2 (출력 레이어)
    outputs = tf.keras.layers.Dense(10, activation='softmax')(x)

    # 모델 생성
    model = tf.keras.Model(inputs=inputs, outputs=outputs)

    return model

# 8. 모델 생성 및 컴파일
# 모델 생성 및 컴파일
model = create_cnn_model()
model.compile(
    optimizer=tf.keras.optimizers.Adam(learning_rate=1e-6),
    loss='categorical_crossentropy',
    metrics=['accuracy']
)
model.summary()

# 9. 배치 크기 설정
batch_size = 16  # 원래 배치 크기 유지

# 커스텀 데이터셋 생성
train_dataset = tf.data.Dataset.from_generator(
    lambda: image_generator(training_sheet, IMAGESAVEURL_training, batch_size, True),
    output_signature=(
        tf.TensorSpec(shape=(None, 64, 64, 1), dtype=tf.float32),
        tf.TensorSpec(shape=(None, 10), dtype=tf.float32)
    )
).prefetch(tf.data.AUTOTUNE)

test_dataset = tf.data.Dataset.from_generator(
    lambda: image_generator(validation_sheet, IMAGESAVEURL_validation, batch_size, False),
    output_signature=(
        tf.TensorSpec(shape=(None, 64, 64, 1), dtype=tf.float32),
        tf.TensorSpec(shape=(None, 10), dtype=tf.float32)
    )
).prefetch(tf.data.AUTOTUNE)

# 10. 모델 학습
# 모델 학습을 위한 설정
epochs = 12
steps_per_epoch = train_num // batch_size
validation_steps = test_num // batch_size

# 체크포인트 콜백 추가 (학습 중단 시 저장된 지점부터 다시 시작 가능)
checkpoint_callback = tf.keras.callbacks.ModelCheckpoint(
    filepath='checkpoint_model.weights.h5',  # .weights.h5로 수정
    save_weights_only=True,
    save_best_only=False,
    save_freq='epoch'
)

# 모델 학습
print(f"학습 시작: {epochs} 에포크, 배치 크기 {batch_size}")
try:
    history = model.fit(
        train_dataset,
        epochs=epochs,
        steps_per_epoch=steps_per_epoch,
        validation_data=test_dataset,
        validation_steps=validation_steps,
        callbacks=[checkpoint_callback],
        verbose=1
    )
    print("모델 학습 완료!")
except Exception as e:
    print(f"학습 중 오류 발생: {e}")

# 11. 모델 평가
# 테스트 데이터로 모델 평가
try:
    print("모델 평가 중...")
    test_loss, test_accuracy = model.evaluate(test_dataset, steps=validation_steps)
    print(f"테스트 정확도: {test_accuracy:.4f}")
except Exception as e:
    print(f"평가 중 오류 발생: {e}")

# 12. 모델 저장
# 모델 저장
try:
    model.save('cnn_model.h5')
    print("모델이 cnn_model.h5로 저장되었습니다.")
except Exception as e:
    print(f"모델 저장 중 오류 발생: {e}")

##############appendix###################
# 추가적인 코드 (선택적 실행)
"""
# 필요한 라이브러리 설치
# pip install matplotlib scikit-learn seaborn

import numpy as np
import matplotlib.pyplot as plt
import seaborn as sns
from sklearn.metrics import confusion_matrix

# 예측을 위한 별도의 제너레이터 생성 (혼동 행렬용)
pred_samples = 1000  # 예측에 사용할 샘플 수 제한
pred_batch_size = 50

# 예측 데이터 수집
all_true_labels = []
all_predictions = []

pred_gen = image_generator(validation_sheet, IMAGESAVEURL_validation, pred_batch_size, False)
for i in range(pred_samples // pred_batch_size):
    x_batch, y_batch = next(pred_gen)
    batch_preds = model.predict(x_batch)

    all_true_labels.append(y_batch)
    all_predictions.append(batch_preds)

# 결과 합치기
all_true_labels = np.vstack(all_true_labels)
all_predictions = np.vstack(all_predictions)

# 1. 실제 라벨 (one-hot encoded → label index 변환)
y_true = np.argmax(all_true_labels, axis=1)  # 실제 라벨

# 2. 예측 결과 (확률값 → 가장 높은 확률의 클래스 선택)
y_pred = np.argmax(all_predictions, axis=1)  # 예측된 라벨

# Confusion Matrix 계산
cm = confusion_matrix(y_true, y_pred)

# 퍼센트 값으로 변환 (0으로 나누는 문제 방지)
cm_sum = cm.sum(axis=1, keepdims=True)
cm_percentage = np.divide(cm.astype(float), cm_sum, where=(cm_sum != 0)) * 100  # 0으로 나누는 것 방지
cm_percentage = np.nan_to_num(cm_percentage)  # NaN을 0으로 변환

# 클래스 라벨 정의
labels = ["BF0.18", "BF0.36", "BF0.54", "OF0.18", "OF0.36", "OF0.54", "IF0.18", "IF0.36", "IF0.54", "NO"]

plt.figure(figsize=(8, 6))

# 히트맵 그리기
sns.heatmap(cm_percentage, annot=True, fmt=".2f", cmap="gray_r",
            xticklabels=labels, yticklabels=labels, cbar=True)

# 축 레이블 추가
plt.xlabel("Predict label", fontsize=12)
plt.ylabel("Actual label", fontsize=12)

# 회전된 라벨 스타일 적용
plt.xticks(rotation=45)
plt.yticks(rotation=45)

# 그래프 표시
plt.title("Confusion Matrix (%)", fontsize=14)
plt.tight_layout()
plt.savefig('confusion_matrix.png')
plt.show()
"""
